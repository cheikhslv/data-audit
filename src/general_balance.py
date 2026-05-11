import pandas as pd
import io
import re

EXCEL_EPOCH = pd.Timestamp("1899-12-30")

CLASSES_COMPTES = {
    "1": "Capitaux & Reserves",
    "2": "Immobilisations",
    "3": "Stocks",
    "4": "Tiers (Clients/Fourn.)",
    "5": "Tresorerie",
    "6": "Charges",
    "7": "Produits",
    "9": "Hors bilan",
}


def _lire_xls_xlrd(contenu_bytes):
    import xlrd
    wb = xlrd.open_workbook(file_contents=contenu_bytes)
    result = {}
    for sh in wb.sheets():
        rows = []
        for i in range(sh.nrows):
            row = []
            for j in range(sh.ncols):
                cell = sh.cell(i, j)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append(None)
                else:
                    row.append(cell.value)
            rows.append(row)
        result[sh.name] = rows
    return result


def _lire_xlsx_openpyxl(contenu_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(contenu_bytes), read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        result[name] = rows
    return result


def _est_compte(val):
    s = str(val).strip().replace('.0', '')
    return bool(__import__('re').match(r'^\d{5,8}$', s))


def _classe_compte(num):
    num = str(num).strip()
    if num:
        return CLASSES_COMPTES.get(num[0], "Autre")
    return "Autre"


def _parser_feuille(rows_list):
    if len(rows_list) < 5:
        return pd.DataFrame(), {}
    meta = {"societe": "", "periode": "", "devise": "RWF", "date_debut": "", "date_fin": ""}
    for row in rows_list[:6]:
        if not row or all(v is None for v in row):
            continue
        a = str(row[0] or "").strip()
        if "Soci" in a:
            meta["societe"] = str(row[1] or "").strip()
    comptes = []
    for row in rows_list:
        if not row or not row[0]:
            continue
        if not _est_compte(row[0]):
            continue
        def get(idx, default=0.0):
            v = row[idx] if len(row) > idx and row[idx] is not None else default
            try:
                return float(v)
            except:
                return default
        num_compte = str(row[0]).strip().replace('.0', '')
        description = str(row[2] or "").strip() if len(row) > 2 else ""
        compte = {
            "num_compte": num_compte,
            "description": description,
            "classe": _classe_compte(num_compte),
            "n1_debit": get(3), "n1_credit": get(4), "n1_solde": get(3)-get(4),
            "mvt_debit": get(5), "mvt_credit": get(6), "mvt_net": get(5)-get(6),
            "solde_debit": get(7), "solde_credit": get(8), "solde_net": get(9),
        }
        comptes.append(compte)
    if not comptes:
        return pd.DataFrame(), meta
    df = pd.DataFrame(comptes)
    for col in ["n1_debit","n1_credit","n1_solde","mvt_debit","mvt_credit","mvt_net","solde_debit","solde_credit","solde_net"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df.reset_index(drop=True), meta


def charger_general_balance(fichier):
    nom = fichier.name.lower()
    contenu = fichier.read()
    if nom.endswith(".xls") and not nom.endswith(".xlsx"):
        sheets = _lire_xls_xlrd(contenu)
    else:
        sheets = _lire_xlsx_openpyxl(contenu)
    resultats = {}
    for nom_feuille, rows_list in sheets.items():
        if nom_feuille in ("", "Sheet1", "Sheet2", "Sheet3", "English Dict"):
            continue
        df, meta = _parser_feuille(rows_list)
        if not df.empty:
            resultats[nom_feuille] = {"comptes": df, "meta": meta}
    resultats["annees"] = sorted(k for k in resultats.keys() if k != "annees")
    return resultats


def kpi_general(df):
    total_solde_debit = df["solde_debit"].sum()
    total_solde_credit = df["solde_credit"].sum()
    return {
        "nb_comptes": len(df),
        "nb_actifs": len(df[df["solde_debit"] > 0]),
        "nb_passifs": len(df[df["solde_credit"] > 0]),
        "total_mvt_debit": df["mvt_debit"].sum(),
        "total_mvt_credit": df["mvt_credit"].sum(),
        "total_solde_debit": total_solde_debit,
        "total_solde_credit": total_solde_credit,
        "equilibre": abs(total_solde_debit - total_solde_credit) < 1000,
        "ecart_equilibre": total_solde_debit - total_solde_credit,
    }


def balance_par_classe(df):
    grp = df.groupby("classe").agg(
        nb_comptes=("num_compte","count"),
        mvt_debit=("mvt_debit","sum"),
        mvt_credit=("mvt_credit","sum"),
        solde_debit=("solde_debit","sum"),
        solde_credit=("solde_credit","sum"),
        solde_net=("solde_net","sum"),
    ).reset_index()
    grp["mvt_net"] = grp["mvt_debit"] - grp["mvt_credit"]
    return grp.sort_values("solde_net", key=abs, ascending=False)


def top_comptes_mouvement(df, n=15):
    df = df.copy()
    df["mvt_total"] = df["mvt_debit"] + df["mvt_credit"]
    return df.sort_values("mvt_total", ascending=False).head(n)[
        ["num_compte","description","classe","mvt_debit","mvt_credit","mvt_net","solde_net"]
    ]


def flag_comptes_desequilibres(df, seuil=1.0):
    df = df.copy()
    df["solde_calcule"] = df["solde_debit"] - df["solde_credit"]
    df["ecart"] = (df["solde_net"] - df["solde_calcule"]).abs()
    mask = df["ecart"] > seuil
    return df[mask][["num_compte","description","classe","solde_debit","solde_credit","solde_net","solde_calcule","ecart"]].sort_values("ecart", ascending=False)


def flag_comptes_solde_inhabituel(df):
    flags = []
    c = df[(df["num_compte"].str.startswith("6")) & (df["solde_net"] < 0)].copy()
    c["anomalie"] = "Charge a solde crediteur"
    flags.append(c)
    p = df[(df["num_compte"].str.startswith("7")) & (df["solde_net"] > 0)].copy()
    p["anomalie"] = "Produit a solde debiteur"
    flags.append(p)
    t = df[(df["num_compte"].str.startswith("5")) & (df["solde_credit"] > 0) & (df["solde_debit"] == 0)].copy()
    t["anomalie"] = "Tresorerie a solde crediteur"
    flags.append(t)
    valid = [f for f in flags if not f.empty]
    if not valid:
        return pd.DataFrame()
    result = pd.concat(valid, ignore_index=True)
    cols = ["num_compte","description","classe","solde_debit","solde_credit","solde_net","anomalie"]
    return result[[c for c in cols if c in result.columns]]


def flag_comptes_sans_mouvement(df):
    mask = (df["n1_solde"].abs() > 0) & (df["mvt_debit"] == 0) & (df["mvt_credit"] == 0)
    return df[mask][["num_compte","description","classe","n1_solde","solde_net"]].sort_values("n1_solde", key=abs, ascending=False)


def flag_variation_significative(df_ref, df_cmp, annee_ref, annee_cmp, seuil_pct=50.0):
    ref = df_ref[["num_compte","description","solde_net"]].rename(columns={"solde_net": f"solde_{annee_ref}"})
    cmp = df_cmp[["num_compte","solde_net"]].rename(columns={"solde_net": f"solde_{annee_cmp}"})
    merged = ref.merge(cmp, on="num_compte", how="inner")
    merged["variation_pct"] = (
        (merged[f"solde_{annee_cmp}"] - merged[f"solde_{annee_ref}"]) /
        merged[f"solde_{annee_ref}"].replace(0, pd.NA).abs() * 100
    ).fillna(0).round(1)
    mask = merged["variation_pct"].abs() > seuil_pct
    return merged[mask].sort_values("variation_pct", key=abs, ascending=False)


def resume_flags_general(df):
    return {
        "nb_desequilibres": len(flag_comptes_desequilibres(df)),
        "nb_soldes_inhabituels": len(flag_comptes_solde_inhabituel(df)),
        "nb_sans_mouvement": len(flag_comptes_sans_mouvement(df)),
    }
