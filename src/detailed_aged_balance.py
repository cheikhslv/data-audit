import pandas as pd
import io
import re

EXCEL_EPOCH = pd.Timestamp("1899-12-30")
TYPES_FACTURES  = {"SAINV", "JEMIS", "OPMTC"}
TYPES_PAIEMENTS = {"APPAY"}
TYPES_AVOIRS    = {"REVAO", "SACRN"}


def _lire_xls_xlrd(contenu_bytes):
    import xlrd
    wb = xlrd.open_workbook(file_contents=contenu_bytes)
    result = {}
    for sh in wb.sheets():
        rows = []
        for i in range(sh.nrows):
            row = []
            for j in range(sh.ncols):
                cell = sh.cell(i, j)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(xlrd.xldate_as_datetime(cell.value, wb.datemode))
                elif cell.ctype == xlrd.XL_CELL_EMPTY:
                    row.append(None)
                else:
                    row.append(cell.value)
            rows.append(row)
        result[sh.name] = rows
    return result


def _lire_xlsx_openpyxl(contenu_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(contenu_bytes), read_only=True, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        result[name] = rows
    return result


def _serial_to_date(val):
    try:
        n = float(val)
        if 30000 < n < 70000:
            return EXCEL_EPOCH + pd.Timedelta(days=int(n))
        return pd.NaT
    except Exception:
        return pd.NaT


def _est_ligne_client(row):
    a_val = str(row[0]).strip() if row[0] is not None else ""
    b_val = row[1] if len(row) > 1 else None
    m = re.match(r'^(\d{5,})\s+(.+)$', a_val)
    if m and isinstance(b_val, (int, float)):
        return True, m.group(1), m.group(2).strip()
    return False, "", ""


def _est_ligne_transaction(row):
    if len(row) < 2:
        return False
    b_val = str(row[1]).strip() if row[1] is not None else ""
    return b_val in (TYPES_FACTURES | TYPES_PAIEMENTS | TYPES_AVOIRS)


def _est_ligne_total(row):
    a_val = str(row[0]).strip() if row[0] is not None else ""
    return a_val.lower() in ("total", "solde")


def _parser_feuille(rows_list):
    if len(rows_list) < 5:
        return pd.DataFrame(), pd.DataFrame(), {}

    meta = {"societe": "", "date_ref": None, "date_ref_str": "N/A"}
    for row in rows_list[:5]:
        if not row or all(v is None for v in row):
            continue
        a = str(row[0] or "").strip()
        if "Soci" in a:
            meta["societe"] = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if len(row) > 6 and row[6]:
            if isinstance(row[6], (float, int)) and row[6] > 30000:
                meta["date_ref"] = _serial_to_date(row[6])
                meta["date_ref_str"] = meta["date_ref"].strftime("%d/%m/%Y") if meta["date_ref"] else "N/A"
            elif hasattr(row[6], 'strftime'):
                meta["date_ref"] = pd.Timestamp(row[6])
                meta["date_ref_str"] = meta["date_ref"].strftime("%d/%m/%Y")

    nom_map = {}
    for row in rows_list:
        if not row or not row[0]:
            continue
        is_client, code, nom = _est_ligne_client(row)
        if is_client and code and nom:
            nom_map[code] = nom

    transactions = []
    for row in rows_list:
        if not row or not row[0]:
            continue
        if _est_ligne_total(row):
            continue
        if not _est_ligne_transaction(row):
            continue

        def get(idx, default=None):
            return row[idx] if len(row) > idx and row[idx] is not None else default

        tiers   = str(get(0, "")).strip()
        type_tx = str(get(1, "")).strip()

        echeance_raw = get(3)
        echeance = pd.NaT
        if isinstance(echeance_raw, (float, int)) and echeance_raw > 30000:
            echeance = _serial_to_date(echeance_raw)
        elif hasattr(echeance_raw, 'strftime'):
            echeance = pd.Timestamp(echeance_raw)

        tx = {
            "tiers":      tiers,
            "nom_client": nom_map.get(tiers, tiers),
            "type_tx":    type_tx,
            "reference":  str(get(2, "")).strip(),
            "echeance":   echeance,
            "jours":      float(get(4, 0) or 0),
            "j60_plus":   float(get(5, 0) or 0),
            "j30_59":     float(get(6, 0) or 0),
            "j0_29":      float(get(7, 0) or 0),
            "j_30_m1":    float(get(8, 0) or 0),
            "j_60_m31":   float(get(9, 0) or 0),
            "j_inf_60":   float(get(10, 0) or 0),
            "categorie":  (
                "facture"  if type_tx in TYPES_FACTURES  else
                "paiement" if type_tx in TYPES_PAIEMENTS else
                "avoir"
            ),
        }
        tx["montant"] = (tx["j60_plus"] + tx["j30_59"] + tx["j0_29"] +
                         tx["j_30_m1"] + tx["j_60_m31"] + tx["j_inf_60"])
        transactions.append(tx)

    if not transactions:
        return pd.DataFrame(), pd.DataFrame(), meta

    df_tx = pd.DataFrame(transactions)

    df_clients = df_tx.groupby(["tiers", "nom_client"]).agg(
        nb_transactions = ("reference", "count"),
        nb_factures     = ("categorie", lambda x: (x == "facture").sum()),
        nb_paiements    = ("categorie", lambda x: (x == "paiement").sum()),
        nb_avoirs       = ("categorie", lambda x: (x == "avoir").sum()),
        j60_plus        = ("j60_plus", "sum"),
        j30_59          = ("j30_59",   "sum"),
        j0_29           = ("j0_29",    "sum"),
        j_30_m1         = ("j_30_m1",  "sum"),
        j_60_m31        = ("j_60_m31", "sum"),
        j_inf_60        = ("j_inf_60", "sum"),
        montant_total   = ("montant",  "sum"),
    ).reset_index()

    df_clients["total_overdue"] = (
        df_clients["j60_plus"] + df_clients["j30_59"] + df_clients["j0_29"]
    )
    df_clients["pct_overdue"] = (
        df_clients["total_overdue"] / df_clients["montant_total"].replace(0, pd.NA) * 100
    ).fillna(0).round(1)

    return df_tx, df_clients, meta


def charger_detailed_aged(fichier):
    nom     = fichier.name.lower()
    contenu = fichier.read()
    if nom.endswith(".xls") and not nom.endswith(".xlsx"):
        sheets = _lire_xls_xlrd(contenu)
    else:
        sheets = _lire_xlsx_openpyxl(contenu)
    resultats = {}
    for nom_feuille, rows_list in sheets.items():
        if nom_feuille in ("", "Sheet1", "Sheet2", "Sheet3", "English Dict"):
            continue
        df_tx, df_clients, meta = _parser_feuille(rows_list)
        if not df_tx.empty:
            resultats[nom_feuille] = {"transactions": df_tx, "clients": df_clients, "meta": meta}
    resultats["annees"] = sorted(k for k in resultats.keys() if k != "annees")
    return resultats


def kpi_detailed(df_tx, df_clients):
    return {
        "nb_clients":      len(df_clients),
        "nb_transactions": len(df_tx),
        "nb_factures":     int((df_tx["categorie"] == "facture").sum()),
        "nb_paiements":    int((df_tx["categorie"] == "paiement").sum()),
        "nb_avoirs":       int((df_tx["categorie"] == "avoir").sum()),
        "total_j60_plus":  df_tx["j60_plus"].sum(),
        "total_j30_59":    df_tx["j30_59"].sum(),
        "total_j0_29":     df_tx["j0_29"].sum(),
        "total_non_echu":  df_tx[["j_30_m1","j_60_m31","j_inf_60"]].sum().sum(),
        "total_overdue":   df_tx[["j60_plus","j30_59","j0_29"]].sum().sum(),
        "balance_nette":   df_tx["montant"].sum(),
    }


def top_clients_overdue(df_clients, n=10):
    return df_clients.sort_values("j60_plus", ascending=False).head(n)


def transactions_par_type(df_tx):
    return df_tx.groupby("type_tx").agg(
        nb=("reference", "count"),
        montant=("montant", "sum"),
    ).reset_index().sort_values("nb", ascending=False)


def flag_factures_impayees_60j(df_tx):
    mask = (df_tx["type_tx"] == "SAINV") & (df_tx["j60_plus"] > 0)
    cols = ["tiers", "nom_client", "reference", "echeance", "jours", "j60_plus", "montant"]
    cols = [c for c in cols if c in df_tx.columns]
    return df_tx[mask][cols].sort_values("j60_plus", ascending=False)


def flag_paiements_sans_facture(df_clients):
    mask = (df_clients["nb_paiements"] > 0) & (df_clients["nb_factures"] == 0)
    return df_clients[mask].sort_values("nb_paiements", ascending=False)


def flag_solde_negatif(df_clients):
    mask = df_clients["montant_total"] < 0
    cols = ["tiers", "nom_client", "montant_total", "nb_factures", "nb_paiements", "nb_avoirs"]
    cols = [c for c in cols if c in df_clients.columns]
    return df_clients[mask][cols].sort_values("montant_total")


def flag_echeances_depassees(df_tx, seuil_jours=90):
    mask = (df_tx["categorie"] == "facture") & (df_tx["jours"] >= seuil_jours) & (df_tx["montant"] > 0)
    cols = ["tiers", "nom_client", "reference", "echeance", "jours", "montant", "j60_plus"]
    cols = [c for c in cols if c in df_tx.columns]
    return df_tx[mask][cols].sort_values("jours", ascending=False)


def resume_flags_detailed(df_tx, df_clients):
    return {
        "nb_factures_60j_plus":      len(flag_factures_impayees_60j(df_tx)),
        "nb_paiements_sans_facture": len(flag_paiements_sans_facture(df_clients)),
        "nb_soldes_negatifs":        len(flag_solde_negatif(df_clients)),
        "nb_echeances_90j_plus":     len(flag_echeances_depassees(df_tx, 90)),
    }
